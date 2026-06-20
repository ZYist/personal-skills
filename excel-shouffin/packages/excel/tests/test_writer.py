import sys

import pytest
from openpyxl import Workbook, load_workbook

from writer import convert_value, create_sheet, main


def _fresh_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


# --- convert_value ---


def test_convert_value_none():
    assert convert_value(None) is None


def test_convert_value_bool_passthrough():
    assert convert_value(True) is True
    assert convert_value(False) is False


def test_convert_value_native_numbers():
    assert convert_value(42) == 42
    assert convert_value(3.14) == 3.14


def test_convert_value_numeric_string_to_number():
    assert convert_value("25") == 25
    assert convert_value("3.14") == 3.14


def test_convert_value_plain_text_untouched():
    assert convert_value("hello") == "hello"


def test_convert_value_leading_zero_lost_by_default():
    # 默认行为：纯数字串被强转，前导零丢失（这是 text_columns 要解决的坑）
    assert convert_value("007") == 7


def test_convert_value_force_text_preserves_leading_zero():
    assert convert_value("007", force_text=True) == "007"
    assert convert_value("001", force_text=True) == "001"


def test_convert_value_force_text_coerces_non_string():
    assert convert_value(42, force_text=True) == "42"
    assert convert_value(None, force_text=True) is None


# --- create_sheet ---


def test_create_sheet_basic_headers_and_data():
    wb = _fresh_workbook()
    ws = create_sheet(
        wb,
        {"name": "测试", "headers": ["A", "B"], "data": [["x", 1], ["y", 2]]},
    )
    assert ws.title == "测试"
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=2, column=1).value == "x"
    assert ws.cell(row=2, column=2).value == 1


def test_create_sheet_text_columns_preserves_leading_zero():
    wb = _fresh_workbook()
    ws = create_sheet(
        wb,
        {
            "headers": ["工号", "姓名"],
            "data": [["007", "张三"], ["001", "李四"]],
            "text_columns": [1],  # 1-based：第 1 列（工号）强制文本
        },
    )
    assert ws.cell(row=2, column=1).value == "007"
    assert ws.cell(row=3, column=1).value == "001"
    assert ws.cell(row=2, column=2).value == "张三"
    # 未声明的列仍走默认转换
    assert ws.cell(row=2, column=2).data_type != "n"


def test_create_sheet_without_text_columns_still_converts():
    wb = _fresh_workbook()
    ws = create_sheet(wb, {"headers": ["工号"], "data": [["007"]]})
    assert ws.cell(row=2, column=1).value == 7


# --- main 错误路径 ---


def test_main_no_args_exits(tmp_path, capsys, monkeypatch):
    monkeypatch.setattr(sys, "argv", ["writer.py"])
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 1


def test_main_missing_input_exits(tmp_path, capsys, monkeypatch):
    monkeypatch.setattr(sys, "argv", ["writer.py", str(tmp_path / "nope.json")])
    with pytest.raises(SystemExit) as exc:
        main()
    assert exc.value.code == 1
    assert "not found" in capsys.readouterr().err


def test_main_invalid_json_exits(tmp_path, monkeypatch):
    bad = tmp_path / "bad.json"
    bad.write_text("{invalid", encoding="utf-8")
    monkeypatch.setattr(sys, "argv", ["writer.py", str(bad)])
    with pytest.raises(SystemExit):
        main()


def test_main_empty_sheets_exits(tmp_path, monkeypatch):
    cfg = tmp_path / "empty.json"
    cfg.write_text('{"sheets": []}', encoding="utf-8")
    monkeypatch.setattr(sys, "argv", ["writer.py", str(cfg)])
    with pytest.raises(SystemExit):
        main()


def test_main_no_valid_sheet_data_exits(tmp_path, monkeypatch):
    cfg = tmp_path / "data.json"
    cfg.write_text('{"sheets": [{"name": "empty"}]}', encoding="utf-8")
    monkeypatch.setattr(sys, "argv", ["writer.py", str(cfg)])
    with pytest.raises(SystemExit):
        main()


def test_main_success_writes_xlsx(tmp_path, monkeypatch, capsys):
    cfg = tmp_path / "data.json"
    out = tmp_path / "out.xlsx"
    cfg.write_text(
        '{"sheets": [{"headers": ["A"], "data": [["1", "007"]], "text_columns": [2]}]}',
        encoding="utf-8",
    )
    monkeypatch.setattr(sys, "argv", ["writer.py", str(cfg), str(out)])
    main()  # 正常返回，不 raise
    assert "Success" in capsys.readouterr().out
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=2, column=1).value == 1  # 第 1 列默认转数字
    assert ws.cell(row=2, column=2).value == "007"  # 第 2 列 text_columns 保留前导零
